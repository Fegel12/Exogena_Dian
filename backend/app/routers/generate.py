# -*- coding: utf-8 -*-
"""Rutas de generación de archivos XML para la DIAN."""
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Balance, GeneratedFile, Tenant, TemplateRule
from app.services.template_engine import generar_formato, exportar_excel

router = APIRouter(prefix="/api", tags=["generacion"])


@router.post("/companies/{tenant_id}/generate")
def generar(tenant_id: int, formato: str = Query("1001"), db: Session = Depends(get_db)):
    bal = (db.query(Balance).filter_by(tenant_id=tenant_id)
           .order_by(Balance.id.desc()).first())
    if not bal:
        raise HTTPException(404, "Esta empresa aún no tiene balances importados.")
    try:
        archivos = generar_formato(bal.id, formato, db)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"format_code": formato, "archivos": archivos}


@router.get("/companies/{tenant_id}/files")
def listar_archivos(tenant_id: int, db: Session = Depends(get_db)):
    if not db.get(Tenant, tenant_id):
        raise HTTPException(404, "Empresa no encontrada.")
    return [
        {"id": f.id, "format_code": f.format_code, "file_name": f.file_name,
         "created_at": str(f.created_at or "")}
        for f in db.query(GeneratedFile).filter_by(tenant_id=tenant_id)
        .order_by(GeneratedFile.id.desc()).limit(100).all()
    ]



@router.get("/files/{file_id}/download")
def descargar(file_id: int, db: Session = Depends(get_db)):
    gf = db.get(GeneratedFile, file_id)
    if not gf:
        raise HTTPException(404, "Archivo no encontrado.")
    return Response(
        content=gf.xml_content.encode("ISO-8859-1"),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{gf.file_name}"'},
    )


@router.post("/companies/{tenant_id}/export-excel")
def exportar_excel_endpoint(tenant_id: int, formato: str = Query("1001"),
                            db: Session = Depends(get_db)):
    bal = (db.query(Balance).filter_by(tenant_id=tenant_id)
           .order_by(Balance.id.desc()).first())
    if not bal:
        raise HTTPException(404, "Esta empresa aún no tiene balances importados.")
    try:
        contenido, nombre_archivo = exportar_excel(bal.id, formato, db)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


# ── Parametrización: mapeo concepto ↔ cuenta con active/inactive ──

@router.get("/template-rules")
def listar_reglas(formato: str = Query("1001"), tenant_id: int = Query(1),
                  db: Session = Depends(get_db)):
    """Devuelve reglas agrupadas por concepto con active/inactive."""
    reglas = db.query(TemplateRule).filter_by(format_code=formato)\
        .order_by(TemplateRule.concepto, TemplateRule.cuentas_desde).all()
    grupos = {}
    for r in reglas:
        key = r.concepto
        if key not in grupos:
            grupos[key] = {"concepto": r.concepto, "concepto_nombre": r.concepto_nombre, "cuentas": []}
        if r.cuentas_desde:
            grupos[key]["cuentas"].append({
                "cuenta": r.cuentas_desde,
                "cuenta_hasta": r.cuentas_hasta if r.cuentas_hasta != r.cuentas_desde else None,
                "active": bool(r.active) if r.active is not None else True,
                "rule_id": r.id,
                "campo_valor": r.campo_valor,
            })
    return sorted(grupos.values(), key=lambda g: g["concepto"])


@router.post("/template-rules")
def crear_regla(payload: dict, db: Session = Depends(get_db)):
    fmt = payload.get("format_code", "1001")
    cpt = int(payload["concepto"])
    cuenta = payload.get("cuenta", "")
    if not cuenta:
        raise HTTPException(400, "Falta el código de cuenta")
    existente = db.query(TemplateRule).filter_by(
        format_code=fmt, concepto=cpt, cuentas_desde=cuenta).first()
    if existente:
        existente.active = True
        db.commit()
        return {"id": existente.id, "ok": True, "msg": "Reactivada"}
    r = TemplateRule(format_code=fmt, concepto=cpt,
        concepto_nombre=payload.get("concepto_nombre", ""),
        cuentas_desde=cuenta, cuentas_hasta=cuenta,
        campo_valor=payload.get("campo_valor", "closing"),
        active=True)
    db.add(r); db.commit()
    return {"id": r.id, "ok": True}


@router.patch("/template-rules/{rule_id}")
def toggle_regla(rule_id: int, payload: dict, db: Session = Depends(get_db)):
    r = db.get(TemplateRule, rule_id)
    if not r: raise HTTPException(404, "No encontrada")
    if "active" in payload:
        r.active = bool(payload["active"])
    db.commit()
    return {"id": r.id, "active": r.active}


@router.delete("/template-rules/{rule_id}")
def eliminar_regla(rule_id: int, db: Session = Depends(get_db)):
    r = db.get(TemplateRule, rule_id)
    if not r: raise HTTPException(404, "No encontrada")
    db.delete(r); db.commit()
    return {"ok": True}


# ── Auto-propuesta ──

@router.post("/template-rules/auto-propose")
def auto_proponer(tenant_id: int = Query(1), formato: str = Query("1001"),
                  db: Session = Depends(get_db)):
    """Genera mapeo automático basado en nombres de cuenta vs concepto."""
    from app.models import BalanceRow
    bal = db.query(Balance).filter_by(tenant_id=tenant_id).order_by(Balance.id.desc()).first()
    if not bal: raise HTTPException(404, "Sin balance")
    rows = db.query(BalanceRow.code, BalanceRow.account_name).filter_by(
        balance_id=bal.id, row_type="account").distinct().all()
    conceptos = db.query(TemplateRule.concepto, TemplateRule.concepto_nombre).filter_by(
        format_code=formato).distinct().all()

    # Reglas de coincidencia: palabra clave en nombre cuenta → concepto
    KEYWORDS = {
        "HONORARI": 5002, "COMISION": 5003, "SERVICIO": 5004,
        "ARREND": 5005, "INTERES": 5006, "RENDIMIENT": 5006,
        "COMPRA": 5007, "ACTIVO": 5008, "PARAFISCAL": 5010,
        "SENA": 5010, "ICBF": 5010, "CAJA": 5010,
        "EPS": 5011, "RIESGO": 5011, "SALUD": 5011,
        "PENSION": 5012, "FONDO": 5012,
        "DONACION": 5013, "IMPUESTO": 5015, "CONSUMO": 5066,
        "VIATICO": 5055, "GASTO": 5016, "COSTO": 5016,
        "DEDUCC": 5016, "TECNIC": 5027, "ASISTENC": 5023,
        "REGALIA": 5024, "EXTERIOR": 5068, "CONSULTOR": 5067,
    }
    creados = 0
    for code, name in rows:
        if not name: continue
        name_upper = name.upper()
        for kw, cpt in KEYWORDS.items():
            if kw in name_upper:
                nombre_cpt = next((c[1] for c in conceptos if c[0] == cpt), "")
                existente = db.query(TemplateRule).filter_by(
                    format_code=formato, concepto=cpt, cuentas_desde=code).first()
                if not existente:
                    db.add(TemplateRule(format_code=formato, concepto=cpt,
                        concepto_nombre=nombre_cpt, cuentas_desde=code,
                        cuentas_hasta=code, campo_valor="closing", active=True))
                    creados += 1
                break

    db.commit()
    return {"creados": creados, "msg": f"Propuesta generada: {creados} nuevas asignaciones"}


# ── IMPORTAR configuración ──

@router.post("/template-rules/import")
async def importar_config(file: UploadFile = File(...), formato: str = Query("1001"),
                          db: Session = Depends(get_db)):
    """Importa mapeo desde Excel/CSV/TXT. Columnas: concepto|nombre|cuenta|active"""
    import csv, io as iomod
    content = await file.read()
    filename = (file.filename or "").lower()
    warnings = []
    creados, actualizados = 0, 0

    rows_data = []
    if filename.endswith(('.csv', '.txt')):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(iomod.StringIO(text))
    elif filename.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(iomod.BytesIO(content), read_only=True)
        ws = wb.active
        reader = ws.iter_rows(values_only=True)
    else:
        raise HTTPException(400, "Formato no soportado. Use .xlsx, .csv, o .txt")

    header = None
    for row in reader:
        if not row or not any(row): continue
        row = [str(c).strip() if c else "" for c in row]
        if row[0].lower() in ("concepto", "dian_concept", "codigo"):
            header = row; continue
        if len(row) < 2: continue
        try:
            cpt = int(row[0])
            cuenta = row[2] if len(row) > 2 else row[1]
            active = row[3].lower() not in ("false", "0", "no", "") if len(row) > 3 else True
            nombre = row[1] if len(row) > 2 and not row[1].isdigit() else ""
            rows_data.append((cpt, nombre, cuenta, active))
        except (ValueError, IndexError):
            warnings.append(f"Fila ignorada: {row[:4]}")
            continue

    for cpt, nombre, cuenta, active in rows_data:
        existente = db.query(TemplateRule).filter_by(
            format_code=formato, concepto=cpt, cuentas_desde=cuenta).first()
        if existente:
            existente.active = active
            if nombre: existente.concepto_nombre = nombre
            actualizados += 1
        else:
            db.add(TemplateRule(format_code=formato, concepto=cpt,
                concepto_nombre=nombre, cuentas_desde=cuenta, cuentas_hasta=cuenta,
                active=active, campo_valor="closing"))
            creados += 1
    db.commit()
    return {"creados": creados, "actualizados": actualizados, "warnings": warnings}


# ── EXPORTAR configuración ──

@router.get("/template-rules/export")
def exportar_config(formato: str = Query("1001"), fmt: str = Query("xlsx"),
                    db: Session = Depends(get_db)):
    """Exporta el mapeo actual a Excel o CSV."""
    reglas = db.query(TemplateRule).filter_by(format_code=formato)\
        .order_by(TemplateRule.concepto, TemplateRule.cuentas_desde).all()

    rows = [(r.concepto, r.concepto_nombre or "", r.cuentas_desde or "",
             "true" if r.active else "false")
            for r in reglas if r.cuentas_desde]

    if fmt == "csv":
        import csv, io as iomod
        out = iomod.StringIO(); w = csv.writer(out)
        w.writerow(["concepto", "concepto_nombre", "puc_account", "active"])
        for r in rows: w.writerow(r)
        out.seek(0)
        return Response(content=out.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=parametrizacion_{formato}.csv"})

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Formato {formato}"
    ws.append(["concepto", "concepto_nombre", "puc_account", "active"])
    for r in rows: ws.append(list(r))
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return Response(content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=parametrizacion_{formato}.xlsx"})


# ── Lista de conceptos DIAN por formato ──

@router.get("/conceptos-dian")
def listar_conceptos(formato: str = Query("1001"), db: Session = Depends(get_db)):
    """Devuelve conceptos con código y nombre para dropdowns."""
    reglas = db.query(TemplateRule).filter_by(format_code=formato)\
        .order_by(TemplateRule.concepto).all()
    return [
        {"codigo": r.concepto, "nombre": r.concepto_nombre,
         "cuentas_desde": r.cuentas_desde, "cuentas_hasta": r.cuentas_hasta}
        for r in reglas
    ]


# ── Lista de cuentas del balance ──

@router.get("/cuentas-balance")
def listar_cuentas_balance(tenant_id: int = Query(1), db: Session = Depends(get_db)):
    """Devuelve cuentas únicas del balance con código y nombre."""
    from app.models import BalanceRow
    bal = db.query(Balance).filter_by(tenant_id=tenant_id)\
        .order_by(Balance.id.desc()).first()
    if not bal:
        return []
    cuentas = db.query(BalanceRow.code, BalanceRow.account_name)\
        .filter_by(balance_id=bal.id, row_type="account")\
        .distinct().order_by(BalanceRow.code).all()
    return [{"codigo": c[0], "nombre": c[1]} for c in cuentas]


# ── Generar TODOS los formatos ──

@router.post("/companies/{tenant_id}/generate-all")
def generar_todos(tenant_id: int, db: Session = Depends(get_db)):
    """Genera XML para todos los formatos que tengan reglas configuradas."""
    from app.services.template_engine import generar_formato
    bal = db.query(Balance).filter_by(tenant_id=tenant_id)\
        .order_by(Balance.id.desc()).first()
    if not bal:
        raise HTTPException(404, "No hay balances importados.")
    
    formatos = [r[0] for r in db.query(TemplateRule.format_code)
                .distinct().order_by(TemplateRule.format_code).all()]
    
    resultados = []
    for fmt in formatos:
        try:
            archivos = generar_formato(bal.id, fmt, db)
            resultados.append({"formato": fmt, "archivos": len(archivos), "ok": True})
        except Exception as e:
            resultados.append({"formato": fmt, "error": str(e), "ok": False})
    
    return {"generados": len([r for r in resultados if r["ok"]]),
            "total_formatos": len(formatos), "detalle": resultados}

